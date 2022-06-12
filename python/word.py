import requests
from bs4 import BeautifulSoup
import json
import time
import xlrd
from openpyxl import load_workbook

DATA = {
    'n':'Noun',
    'v':'Verb',
    'prep':'preposition',
    'adj':'adjective',
    'adv':'adverb'
}

class Word():
    def __init__(self):
        self.url = 'http://cn.bing.com/dict/search?q='
        self.word_info_selector_dict={
            'word':'div#headword',
            'word_warning': 'pass',
            'eng_pr':'div.hd_pr.b_primtxt',
            'ame_pr':'div.hd_prUS.b_primtxt',
            'tongyi':'div.wd_div',
            'fushu':'div.hd_div1',
            'defination':'div.qdef > ul > li',
        }
        self.word_info_dict={}
    def split_translate(self):
        pass


    def get_content(self,from_obj,selector_str):
        rst = from_obj.select(selector_str)
        
        # print(rst)
        if len(rst)==0:
        
            return None
        elif len(rst)==1:
            return rst[0].get_text().replace("\xa0"," ").strip()
        else:
            return '\n'.join([e.get_text() for e in rst])

    def look_up(self,word):
        url = self.url + word
        self.text = requests.get(url,headers={'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36 Edg/94.0.992.31'}).text
        sp = BeautifulSoup(self.text,'html.parser')

        self.related_divs = sp.select('div.lf_area > div')
        # print(url)
        # print(self.related_divs)
        for k,v in self.word_info_selector_dict.items():
            x = self.get_content(self.related_divs[0],v)
            if x == 'pass':
                continue
            if x== None:
                if  k == 'word_warning':
                    x = self.related_divs[0].get_text()

            self.word_info_dict[k]=x
        
        self.word_info_dict['sentences']=self.get_content(self.related_divs[1],'div#sentenceSeg')

        # print(self.word_info_dict)

    def get_def(self):
        def_content = self.word_info_dict['defination']
        if def_content ==  None:
            print('Warning: %s'%self.word_info_dict['word_warning'])
            
            return None
        finished = def_content.split('\n')#['adj.清楚的；明显的；浅白的；坦诚的', 'v.发牢骚；叹惜；哀悼；痛哭', 'adv.绝对地', 'n.平原', '网络普通的；平纹；朴素的']
        self.def_dic = {
                
            }
        for i in finished:
            def_list= i.split('.')#['adj', '清楚的；明显的；浅白的；坦诚的']
            if len(def_list) == 1:
                continue
            if len(def_list[1].split('；')) > 2:
                zh = def_list[1].split('；')[0]+';'+def_list[1].split('；')[1]

                self.def_dic[def_list[0]]=zh
            else:
                self.def_dic[def_list[0]]=def_list[1]
        return True
    def get_pr(self):
        pr_content = self.word_info_dict['eng_pr']
        # print(self.word_info_dict)
        return pr_content


    def get_cha(self,cha):
        cha = cha.replace(' ','').lower()
        x = self.get_def()
        if x ==None:
            return None
        
        # print(self.def_dic[cha])
        if cha in self.def_dic:
            niu = self.def_dic[cha]
            return niu
        elif cha not in DATA:
            print('%s不存在'%self.word_info_dict['word']+cha)
        else:
            
            print('%s不存在'%self.word_info_dict['word']+DATA[cha])
            # print(self.def_dic)
            return None
        
            
        

class Xls():
    def __init__(self):
        self.path = '/Users/bailu/Downloads/单词.xlsx'
        self.workbook = load_workbook(self.path)
        self.sheet = self.workbook['阅读单词']
    def start(self):
        
        xlsx = xlrd.open_workbook(self.path)
        table = xlsx.sheet_by_index(0)
        nrows = table.nrows# 获取表格行数
        word = Word()
        n = 0
        p = 0
        for x in range(1,nrows):
            wd = table.cell_value(x,0)
            pr = table.cell_value(x,3) #音标
            raw_cha = table.cell_value(x,1) #获取词性
            if raw_cha != None:
                cha = raw_cha.split('.')[0]
                
            

            zh = table.cell_value(x,2)
            if type(wd) != float and cha != '' and zh == '':
                if cha == 'pp':
                    cha = 'prep'
                
                if wd == '':
                    continue
                self.danci = word.look_up(wd)
                # self.fanyi =''
                
                y = word.get_cha(cha)
                # print(y)
                if y == None:
                    continue
                self.fanyi = y
                
                
                n = n+1
                self.sheet.cell(x+1, 3).value = self.fanyi
                # print(x)
            if type(wd) != float and wd != '' and pr == '':
                self.danci = word.look_up(wd) #可以去掉
                pr = word.get_pr()
                if pr == '':
                    continue
                print(x+1)
                self.sheet.cell(x+1,4).value=pr
                p = p+1
            # if p ==5:
            #     break

                
                
                
        
        print('共翻译%s个单词'%n)
        print('共添加%s个音标'%p)
        self.workbook.save(self.path)
        self.workbook.close()



                


if __name__ == '__main__':
    xls = Xls()
    xls.start()




